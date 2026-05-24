import hashlib
import fnmatch
import os
from datetime import datetime
from pathlib import Path
from typing import Callable, Dict, List, Optional
from concurrent.futures import ThreadPoolExecutor, as_completed

try:
    import openpyxl
    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False
    console.print("[yellow][预警] openpyxl 未安装，备份记录不会写入 Excel 文件。运行: pip install openpyxl[/]")

from quark_client import QuarkClient
from rich.console import Console
from rich.progress import (
    BarColumn, Progress, SpinnerColumn,
    TextColumn, TimeRemainingColumn, TransferSpeedColumn,
)

from .auth import ensure_remote_dirs
from .config import AppConfig, SourceConfig
from .uploader import QuarkUploader

console = Console()

def _log_to_xls(file_name: str, remote_path: str, log_path: str = "data.xlsx"):
    if not HAS_EXCEL:
        return
    try:
        from openpyxl.styles import Font, PatternFill, Alignment, Side, Border
        from openpyxl.utils import get_column_letter
        lp = Path(log_path)
        # Force .xlsx extension (openpyxl doesn't support old .xls format)
        if lp.suffix.lower() != ".xlsx":
            lp = lp.with_suffix(".xlsx")
        lp.parent.mkdir(parents=True, exist_ok=True)
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        if lp.exists():
            wb = openpyxl.load_workbook(lp)
            ws = wb.active
            row_num = ws.max_row + 1
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "备份记录"
            ws.append(["备份文件名", "备份时间", "远程路径"])
            row_num = 2

            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_align = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style="thin", color="B4C6E7"),
                right=Side(style="thin", color="B4C6E7"),
                top=Side(style="thin", color="B4C6E7"),
                bottom=Side(style="thin", color="B4C6E7"),
            )
            for col in range(1, 4):
                cell = ws.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border

            ws.auto_filter.ref = "A1:C1"
            ws.column_dimensions["A"].width = 40
            ws.column_dimensions["B"].width = 22
            ws.column_dimensions["C"].width = 50

        ws.append([file_name, now, remote_path])
        new_row = ws.max_row

        data_font = Font(size=11)
        data_align = Alignment(vertical="center")
        thin_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )
        alt_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

        for col in range(1, 4):
            cell = ws.cell(row=new_row, column=col)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border
            if new_row % 2 == 0:
                cell.fill = alt_fill

        if ws.auto_filter.ref:
            last_col = get_column_letter(3)
            ws.auto_filter.ref = f"A1:{last_col}{new_row}"

        wb.save(lp)
    except Exception as e:
        console.print(f"[red][错误] 写入备份记录失败: {e}[/]")


def _file_hash(path: Path) -> str:
    h = hashlib.md5()
    with open(path, "rb") as f:
        while chunk := f.read(65536):
            h.update(chunk)
    return h.hexdigest()


def _should_exclude(file_path: Path, source: SourceConfig) -> bool:
    rel = file_path.relative_to(Path(source.local)).as_posix()
    for pattern in source.exclude:
        if fnmatch.fnmatch(rel, pattern) or fnmatch.fnmatch(file_path.name, pattern):
            return True
    return False


def _should_include(file_path: Path, source: SourceConfig) -> bool:
    if not source.include:
        return True
    rel = file_path.relative_to(Path(source.local)).as_posix()
    return any(fnmatch.fnmatch(rel, p) or fnmatch.fnmatch(file_path.name, p) for p in source.include)


def _walk_local_files(source: SourceConfig) -> List[Path]:
    base = Path(source.local)
    if not base.exists():
        return []
    if base.is_file():
        return [base] if (_should_include(base, source) and not _should_exclude(base, source)) else []
    files = []
    if source.recursive:
        for f in base.rglob("*"):
            if f.is_file() and _should_include(f, source) and not _should_exclude(f, source):
                files.append(f)
    else:
        for f in base.iterdir():
            if f.is_file() and _should_include(f, source) and not _should_exclude(f, source):
                files.append(f)
    return files


class QuarkBackup:
    def __init__(self, client: QuarkClient, config: AppConfig):
        self.client = client
        self.config = config
        self.cache: Dict[str, str] = config.load_cache()
        self._cancel = False
        cookie_str = client.api_client.cookies or ""
        self.uploader = QuarkUploader(cookie_str)

    def cancel(self):
        self._cancel = True

    def run_source(self, source: SourceConfig, progress_callback: Optional[Callable] = None) -> Dict:
        local_files = _walk_local_files(source)
        if not local_files:
            if progress_callback:
                progress_callback("status", source.local, "无文件需要备份")
            return {"status": "skipped", "uploaded": 0, "skipped": 0, "failed": 0}

        remote_folder_id = ensure_remote_dirs(self.client, source.remote)

        upload_tasks = []
        for lf in local_files:
            if self._cancel:
                return {"status": "cancelled", "uploaded": 0, "skipped": 0, "failed": 0}
            rel = lf.relative_to(Path(source.local)).as_posix()
            cache_key = f"{source.remote}/{rel}"
            current_md5 = _file_hash(lf)
            cached_md5 = self.cache.get(cache_key)
            if cached_md5 == current_md5:
                continue
            upload_tasks.append((lf, rel, source, remote_folder_id))

        if not upload_tasks:
            if progress_callback:
                progress_callback("status", source.local, "全部已是最新")
            return {"status": "up_to_date", "uploaded": 0, "skipped": len(local_files), "failed": 0}

        total = len(upload_tasks)
        uploaded = 0
        skipped = 0
        failed = 0

        if progress_callback:
            progress_callback("begin", source.local, {"total": total, "remote": source.remote})
        else:
            console.print(f"  [cyan]需上传 {total} 个文件[/]")

        if progress_callback is None:
            progress_ctx = Progress(
                SpinnerColumn(),
                TextColumn("[progress.description]{task.description}"),
                BarColumn(),
                TextColumn("[progress.percentage]{task.percentage:>3.0f}%"),
                TransferSpeedColumn(),
                TimeRemainingColumn(),
                console=console,
            )
            progress_ctx.__enter__()
            task = progress_ctx.add_task(f"[cyan]上传中...", total=total)

        max_workers = self.config.concurrency.max_upload_workers
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = {}
            for lf, rel, src, rfi in upload_tasks:
                future = executor.submit(self._upload_single, lf, rfi, rel)
                futures[future] = (lf, rel)

            for future in as_completed(futures):
                if self._cancel:
                    break
                lf, rel = futures[future]
                try:
                    ok = future.result()
                    if ok:
                        cache_key = f"{source.remote}/{rel}"
                        self.cache[cache_key] = _file_hash(lf)
                        uploaded += 1
                        _log_to_xls(lf.name, cache_key, self.config.backup_log_path)
                        if source.delete_after_backup:
                            try:
                                lf.unlink()
                                if progress_callback:
                                    progress_callback("status", f"已删除: {lf.name}", None)
                            except Exception as del_e:
                                if progress_callback:
                                    progress_callback("status", f"删除失败: {lf.name} - {del_e}", None)
                    else:
                        failed += 1
                except Exception as e:
                    failed += 1
                    if progress_callback:
                        progress_callback("file_fail", f"{lf.name}: {e}", None)
                if progress_callback:
                    progress_callback("file_done", lf.name, {"uploaded": uploaded, "failed": failed, "total": total})
                else:
                    progress_ctx.update(task, advance=1)

        if progress_callback is None:
            progress_ctx.__exit__(None, None, None)

        self.config.save_cache(self.cache)
        result = {"status": "completed", "uploaded": uploaded, "skipped": skipped, "failed": failed}
        if progress_callback:
            progress_callback("done", source.local, result)
        return result

    def _upload_single(self, local_path: Path, parent_fid: str, rel: str) -> bool:
        result = self.uploader.upload_file(
            file_path=str(local_path),
            parent_folder_id=parent_fid,
        )
        return result.get("status") == "success"

    def run_all(self, progress_callback: Optional[Callable] = None) -> List[Dict]:
        results = []
        for source in self.config.sources:
            if self._cancel:
                break
            if progress_callback is None:
                console.rule(f"[bold]备份: {source.local} → {source.remote}")
            try:
                result = self.run_source(source, progress_callback)
                results.append({"source": source.local, **result})
            except Exception as e:
                if progress_callback is None:
                    console.print(f"[red]备份失败: {e}[/]")
                results.append({"source": source.local, "status": "error", "error": str(e)})
        self._cancel = False
        return results
